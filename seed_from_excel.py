import os
from openpyxl import load_workbook
from app import create_app, db
from app.models import Item, Ability, Condition, Display, User, Character, GameState
from werkzeug.security import generate_password_hash
from datetime import datetime, timezone

app = create_app()

def get_cell(row, i):
    return str(row[i].value).strip() if row[i].value else ""

def populate_from_excel(filename):
    wb = load_workbook(filename)
    with app.app_context():
        db.drop_all()
        db.create_all()

        # Items
        if 'Items' in wb.sheetnames:
            for row in list(wb['Items'].iter_rows(min_row=2)):
                db.session.add(Item(
                    name=get_cell(row, 0),
                    description=get_cell(row, 1),
                    code=get_cell(row, 2) or None
                ))

        # Abilities
        if 'Abilities' in wb.sheetnames:
            for row in list(wb['Abilities'].iter_rows(min_row=2)):
                db.session.add(Ability(
                    name=get_cell(row, 0),
                    description=get_cell(row, 1)
                ))

        # Conditions
        if 'Conditions' in wb.sheetnames:
            for row in list(wb['Conditions'].iter_rows(min_row=2)):
                raw_clear_method = get_cell(row, 2)  # or whatever column index
                clear_method = [a.strip() for a in raw_clear_method.split(',') if a.strip()]
                raw_effect = get_cell(row, 3)  # or whatever column index
                effect = [a.strip() for a in raw_effect.split(',') if a.strip()]

                db.session.add(Condition(
                    name=get_cell(row, 0),
                    description=get_cell(row, 1),
                    clear_method=clear_method,
                    effect=effect
                ))

        # Displays
        if 'Displays' in wb.sheetnames:
            for row in list(wb['Displays'].iter_rows(min_row=2)):
                db.session.add(Display(
                    name=get_cell(row, 0),
                    location=get_cell(row, 1),
                    message=get_cell(row, 2),
                    alert_level=get_cell(row, 3) or 'Nominal',
                    animation_mode=get_cell(row, 4) or 'pulse'
                ))

        # Character Roles
        if 'Characters' in wb.sheetnames:
            for row in list(wb['Characters'].iter_rows(min_row=2)):
                raw_abilities = get_cell(row, 2)  # or whatever column index
                abilities = [a.strip() for a in raw_abilities.split(',') if a.strip()]
                db.session.add(Character(
                    position=get_cell(row, 0),
                    affiliation=get_cell(row, 1),
                    abilities=abilities,
                    items=get_cell(row, 3),
                    claim_code=get_cell(row, 4)
                ))



        # Users
        if 'Users' in wb.sheetnames:
            for row in list(wb['Users'].iter_rows(min_row=2)):
                username = get_cell(row, 0)
                email = get_cell(row, 1)
                raw_password = get_cell(row, 2)
                role = get_cell(row, 3) or 'Player'
                theme = get_cell(row, 4) or 'theme-green'

                if username and email and raw_password:
                    hashed_password = generate_password_hash(raw_password)
                    db.session.add(User(
                        username=username,
                        email=email,
                        password=hashed_password,
                        role=role,
                        theme=theme,
                        is_verified=True
                    ))

        # GameState (only 1 row expected)
        if 'GameState' in wb.sheetnames:
            sheet = wb['GameState']
            for row in list(sheet.iter_rows(min_row=2)):
                id_val = int(row[0].value or 1)
                current_phase = str(row[1].value or "Not Started").strip()
                duration = int(row[2].value or 40)
                alert_level = str(row[3].value or "Nominal").strip()

                existing = GameState.query.get(id_val)
                if existing:
                    existing.current_phase = current_phase
                    existing.phase_duration_minutes = duration
                    existing.alert_level = alert_level
                else:
                    db.session.add(GameState(
                        id=id_val,
                        current_phase=current_phase,
                        phase_duration_minutes=duration,
                        alert_level=alert_level
                    ))


        db.session.commit()
        print("Seed data loaded successfully.")

if __name__ == '__main__':
    populate_from_excel('C:/Users/Matt/Documents/Redshift/seed_data.xlsx')
